import quicky
import sql_models
def do_export(request):
    import os
    from django.http import HttpResponse
    import openpyxl
    from openpyxl import utils
    session = quicky.sql_db.begin_session(request.get_app().settings.get_db_config())
    total_items = 0
    ret = {}
    try:

        total_items = session.query(sql_models.models.AuthUser).count()
        items = list(session.query(sql_models.models.AuthUser))
        quicky.sql_db.end_session(session)
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        header_attr=["username","password","email","first_name","last_name","is_active","is_staff","is_superuser"]
        for x in header_attr:
            col_index=header_attr.index(x)+1
            key =  os.path.dirname(request.get_app().mdl.__file__)
            key =__file__[key.__len__()+1:__file__.__len__()]
            ws.cell(1, col_index).value = request.get_app_res("{0}/{1}".format( key,x),x)
        row_index=2
        for row in items:
            for x in header_attr:
                col_index = header_attr.index(x) + 1
                ws.cell(row_index, col_index).value = (lambda v:  getattr(row,v,None) if v!="password" else None)(x)
            row_index = row_index +1
        col_index =1
        for x in header_attr:
            wb.create_named_range(x, ws, "$"+utils.get_column_letter(col_index)+":$"+utils.get_column_letter(col_index))
            col_index = col_index +1
        from openpyxl.writer.excel import save_virtual_workbook
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename= "{}"'.format("users.xlsx")
        return response


    except Exception as ex:
        quicky.sql_db.end_session(session)
        raise ex
    return ret

    return {}